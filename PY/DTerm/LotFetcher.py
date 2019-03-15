import urllib.request as ulib
import os
import sys
import openpyxl as pyxl
from openpyxl import Workbook
import win32api
import datetime

def returnargs():
    arguments = []
    try:
        for argument in sys.argv[1:]:
            print(argument)
            arguments.append(argument)
    except:
        pass
    
    return arguments

def lineparse(line):

    firstSweep = line.split("<")
    secondSweep = firstSweep[1].split(">")

    num = secondSweep[1]

    return num


def main():

    print("Bienvenido al importador de loterias de porodo!")

    arguments = returnargs()
    if arguments:
        choosing = True
    else:
        choosing = False

    url = "https://resultadodelaloteria.com/colombia"

    loterias = {
        "Cu/marca": "loteria-de-cundinamarca", 
        "Cruz Roja":"loteria-de-la-cruz-roja",
        "Manizales":"loteria-de-manizales",
        "Bogota":"loteria-de-bogota",
        "Boyaca":"loteria-de-boyaca",
        "S/tander":"loteria-de-santander"}
    
    if not choosing:

        wb = Workbook()
        ws = wb.active
        ws.page_margins.left = 0
        ws.page_margins.right = 0
        ws.page_margins.top = 0
        ws.page_margins.bottom = 0

        now = datetime.datetime.now()
        month = now.strftime("%B")

        fecha = f"{now.year}/{month}//{now.day}"
        ws["A1"].value = fecha
        ws["A3"].value = "NUMEROS"

        index = 4
        for loteria in loterias:

            loturl = f"{url}/{loterias[loteria]}"
            print(loturl)
            html = ulib.urlopen(loturl).readlines()

            for line in html:
                if "ctl00_ContentPlaceHolder1_lbl4Cifras" in str(line):
                    break

            print("Extrayendo numero . . .")

            num = lineparse(str(line))
            print(f"El numero es: {num}!")

            ws[f"A{index}"].value = f"LOT {loteria}: [{num}]"
            rpath = f"{os.path.dirname(__file__)}\\temp.xlsx"
            
            index += 1

    wb.save(rpath)

    printer = "POS-58"

    win32api.ShellExecute (
            0,
            "printto",
            f"{rpath}",
            f"{printer}",
            ".",
            0
            )
    
    os.remove(rpath)

main()