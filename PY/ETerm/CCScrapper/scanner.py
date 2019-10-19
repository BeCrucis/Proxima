import os
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import openpyxl

nombreExcel = "HojaEjemplo.xlsx"

excelWorkbook = openpyxl.load_workbook(nombreExcel)
excelSheet = excelWorkbook.active

documents = []

for row in excelSheet.iter_rows(1, excelSheet.max_row, 1, 1):
    for cell in row:
        documents.append(cell.value)
# time.sleep(6)

print(documents)

driver = webdriver.Chrome("F:/Downloads/Carpetas/Utility/Temp/chromedriver.exe")

driver.get("https://www.uis.edu.co/sistemasInformacion/login.seam?josso_back_to=https://www.uis.edu.co/sistemasInformacion/josso_security_check")

a = "636172746172617a"
b = "73616e746931343732"

username = bytearray.fromhex(a).decode()
password = bytearray.fromhex(b).decode()

userTextbox = driver.find_elements_by_id("josso_username")[0]
userTextbox.send_keys(username)

passTextbox = driver.find_elements_by_id("josso_password")[0]
passTextbox.send_keys(password)

enterButton = driver.find_elements_by_id("boton")[0]
enterButton.click()

os.system("pause")

found = False

while not found:
    try:
        tipoDocumento = driver.find_elements_by_id("form1:dcrDocumentoIdentificacion:txtDocumentoIdentificacion")[0]
        tipoDocumento.send_keys("prueba")
        found = True
    except:
        time.sleep(1)

validatedURL = driver.current_url

loopIndex = 0
for document in documents:

    if not document:
        continue 

    loopIndex += 1

    driver.get(validatedURL)
    tipoDocumento = driver.find_elements_by_id("form1:dcrDocumentoIdentificacion:txtDocumentoIdentificacion")[0]
    tipoDocumento.send_keys(document)

    botonConsultar = driver.find_elements_by_id("form1:btnConsultar")[0]
    botonConsultar.click()
    time.sleep(0.5)

    try:
        botonConsultarEstudiante = driver.find_elements_by_id("form1:j_id316:0:btnVerDetalle")[0]
        botonConsultarEstudiante.click()
        time.sleep(0.5)
    except:
        print(F"La cedula {document} no tiene estudiante asociado")
        continue

    try:
        linkInfo = driver.find_elements_by_link_text("Programas Académicos")[0]
        linkInfo.click()
    except:
        print(F"Error en cedula {document}, reintentando . . .")

        try:
            time.sleep(2)
            linkInfo = driver.find_elements_by_link_text("Programas Académicos")[0]
            linkInfo.click()
        except:
            print(F"Error en cedula {document}, reintento fallido, continuando . . .")
            continue

    studentCareer = driver.find_elements_by_id("j_id182:j_id232:0:j_id239")[0].text
    studentCareer = ''.join([i for i in studentCareer if not i.isdigit()]).strip()

    careerQuantity = driver.find_elements_by_id("j_id182:totalRegistros")[0].text
    if "1" not in careerQuantity:
        studentCareer = "VARIAS CARRERAS DETECTADAS - COLOCAR MANUALMENTE"

    print(F"La cedula {document} tiene la carrera de {studentCareer}")

    excelSheet.cell(loopIndex, 2).value = studentCareer

    if (loopIndex % 10) == 0:
        excelWorkbook.save(F"Output{loopIndex}.xlsx")

excelWorkbook.save("FinalOutput.xlsx")

