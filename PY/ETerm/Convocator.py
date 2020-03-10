import openpyxl
import unidecode
import os

wb = openpyxl.load_workbook("Grupos_2020.xlsx")

announcement_ws = wb["Facultad de Ciencias"]
due_teachers_ws = wb["Profesores_Mora"]

# Obtiene la lista de los profesores en mora
# La lista contendra listas con las palabras individuales de los profesores

due_teachers_column = 1
due_teachers_start_row = 4
due_teachers_names = []

# Recorre las filas con las celdas con los nombres de los profesores
for row in due_teachers_ws.iter_rows(min_row=due_teachers_start_row, 
                                    max_row=due_teachers_ws.max_row,
                                    min_col=1,
                                    max_col=1):
    # Extrae la celda de la fila
    due_teacher_cell = row[0]
    # Verifica que la celda no este  vacia
    if not due_teacher_cell.value or not due_teacher_cell.value.strip():
        continue
    # Extrae el nombre tal cual esta en la celda y le quita los espacios laterales
    due_teacher_raw_name = due_teacher_cell.value.strip()
    # Extrae las palabras individuales del nombre
    due_teacher_name = due_teacher_raw_name.split(" ")
    temp = due_teacher_name
    # Sanitiza los nombres para sacar los espacios sobrantes
    due_teacher_name = [name.strip() for name in temp if name.strip()]
    
    # Sanitiza los nombres para sacar los acentos
    for index,word in enumerate(due_teacher_name):

        due_teacher_name[index] = unidecode.unidecode(word)
    
    due_teachers_names.append(due_teacher_name)

# Ahora verifica los nombres de los directores de grupo

leader_teachers_column = 7
leader_teachers_start_row = 5

# Recorre las filas con las celdas con los nombres de los profesores
for row in announcement_ws.iter_rows(min_row=leader_teachers_start_row, 
                                    max_row=announcement_ws.max_row, 
                                    min_col=leader_teachers_column,
                                    max_col=leader_teachers_column):

    # Extrae la celda de la fila
    leader_teacher_cell = row[0]
    # Verifica que la celda no este  vacia
    if not leader_teacher_cell.value or not leader_teacher_cell.value.strip():
        continue
    # Extrae el nombre tal cual esta en la celda y le quita los espacios laterales
    leader_teacher_raw_name = leader_teacher_cell.value.strip()
    # Extrae las palabras individuales del nombre
    leader_teacher_name = leader_teacher_raw_name.split(" ")
    temp = leader_teacher_name
    # Sanitiza los nombres para sacar los espacios sobrantes
    leader_teacher_name = [name.strip() for name in temp if name.strip()]
    
    # Sanitiza los nombres para sacar los acentos
    for index,word in enumerate(leader_teacher_name):

        leader_teacher_name[index] = unidecode.unidecode(word)
    
    # Compara nombres entre el nombre del lider y la lista de nombres de profesores en mora
    is_due = False

    for name in due_teachers_names:
        # Busca palabras en comun entre nombres
        word_ocurrences = len([word for word in leader_teacher_name if word in name])
        if word_ocurrences == len(name):
            # Si hay una equivalencia exacta, probablemente este en mora
            is_due = True
            break
    
    if is_due:
        announcement_ws.cell(leader_teacher_cell.row, leader_teacher_cell.column+2).value= "MORA"
        
wb.save("Output.xlsx")