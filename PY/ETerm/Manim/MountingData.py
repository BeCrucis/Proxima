import openpyxl 

excelWB = openpyxl.load_workbook("Datos i1.xlsx")
excelWS = excelWB.active

STARTROW = 5
STARTCOLUMN = 6
DATASIZE = 5

ELECTRODE_MAX_X = 6
ELECTRODE_MAX_Y = 5
ELECTRODE_RADIUS = 1
CIRCLE_RADIUS = 2

class RR():
    
    startRow = STARTROW
    startColumn = STARTCOLUMN
    finishRow = STARTROW + 4
    finishColumn = STARTCOLUMN + 2*(DATASIZE-1) + 1

    def __init__(self):

        self.equiLines = self.getEquipotentialLines()

        # for line in self.equiLines:
        #     print(F"{line} : {self.equiLines[line]}")

        self.electrodes = self.createElectrodes()
    
    def getEquipotentialLines(self):

        equiLines = {}

        rowIndex = RR.startRow
        
        for row in excelWS.iter_rows(RR.startRow, RR.finishRow, RR.startColumn, RR.finishColumn):
            coords = []
            voltage = int(excelWS.cell(rowIndex, RR.startColumn-1).value)
            equiLines[voltage] = []
            rowIndex += 1 

            cellindex = 0
            for x,y in zip(row[::2], row[1::2]):

                equiLines[voltage].append((round(x.value,4),round(y.value,4)))
        
        return equiLines
    
    def createElectrodes(self):

        leftElectrodeCorners = (
            (-ELECTRODE_MAX_X, -ELECTRODE_MAX_Y),
            (-ELECTRODE_MAX_X-ELECTRODE_RADIUS, -ELECTRODE_MAX_Y),
            (-ELECTRODE_MAX_X-ELECTRODE_RADIUS, ELECTRODE_MAX_Y),
            (-ELECTRODE_MAX_X, ELECTRODE_MAX_Y)
        )

        rightElectrodeCorners = (
            (ELECTRODE_MAX_X, -ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X+ELECTRODE_RADIUS, -ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X+ELECTRODE_RADIUS, ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X, ELECTRODE_MAX_Y)
        )

        return [leftElectrodeCorners, rightElectrodeCorners]

class CR():
    
    startRow = STARTROW + 5
    startColumn = STARTCOLUMN
    finishRow = STARTROW + 4 + 5
    finishColumn = STARTCOLUMN + 2*(DATASIZE-1) + 1

    def __init__(self):

        self.equiLines = self.getEquipotentialLines()
        self.electrodes = self.createElectrodes()
    
    def getEquipotentialLines(self):

        equiLines = {}

        rowIndex = CR.startRow
        
        for row in excelWS.iter_rows(CR.startRow, CR.finishRow, CR.startColumn, CR.finishColumn):
            coords = []
            voltage = int(excelWS.cell(rowIndex, CR.startColumn-1).value)
            equiLines[voltage] = []
            rowIndex += 1 

            cellindex = 0
            for x,y in zip(row[::2], row[1::2]):

                equiLines[voltage].append((round(x.value,4),round(y.value,4)))
        
        return equiLines

    def createElectrodes(self):

        leftElectrodeInfo = (
            (-ELECTRODE_MAX_X - CIRCLE_RADIUS, 0),
            CIRCLE_RADIUS
        )

        rightElectrodeCorners = (
            (ELECTRODE_MAX_X, -ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X+ELECTRODE_RADIUS, -ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X+ELECTRODE_RADIUS, ELECTRODE_MAX_Y),
            (ELECTRODE_MAX_X, ELECTRODE_MAX_Y)
        )

        return [leftElectrodeInfo, rightElectrodeCorners]

class CC():
    
    startRow = STARTROW + 5 + 5
    startColumn = STARTCOLUMN
    finishRow = STARTROW + 5 + 5 + 2
    finishColumn = STARTCOLUMN + 2*(DATASIZE-1) + 1

    def __init__(self):

        self.equiLines = self.getEquipotentialLines()
        self.electrodes = self.createElectrodes()
        
    def getEquipotentialLines(self):

        equiLines = {}

        rowIndex = CC.startRow
        
        for row in excelWS.iter_rows(CC.startRow, CC.finishRow, CC.startColumn, CC.finishColumn):
            coords = []
            voltage = int(excelWS.cell(rowIndex, CC.startColumn-1).value)
            equiLines[voltage] = []
            rowIndex += 1 

            cellindex = 0
            for x,y in zip(row[::2], row[1::2]):

                equiLines[voltage].append((round(x.value,4),round(y.value,4)))
        
        return equiLines
    
    def createElectrodes(self):

        leftElectrodeInfo = (
            (-ELECTRODE_MAX_X - CIRCLE_RADIUS, 0),
            CIRCLE_RADIUS
        )

        rightElectrodeInfo = (
            (ELECTRODE_MAX_X + CIRCLE_RADIUS, 0),
            CIRCLE_RADIUS
        )

        return [leftElectrodeInfo, rightElectrodeInfo]

if __name__ == "__main__":
    RR()