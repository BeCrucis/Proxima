import os
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import time
import openpyxl
from Student import Student
import unidecode
import asyncio
import copy

WD_PATH = 'F:/Downloads/Carpetas/Utility/Temp/chromedriver83.exe'
TERM_NUMBER = 2
TERM_NAME = 'Segundo'
YEAR = '2020'
OED_PAGE = 'https://colmilgeneralsantander.phidias.co/poll/consolidate/people?poll=32'
ED_PAGE = 'https://colmilgeneralsantander.phidias.co/poll/consolidate/people?poll=4'

def click_course(driver, observer_link, course_name):

	driver.get(observer_link)

	course_string_input = "Comportamiento " + course_name

	searchBox = driver.find_elements_by_id("people_search")[0]
	searchBox.send_keys(course_string_input)

	time.sleep(2)

	moreElements = driver.find_elements_by_class_name("result_type_course")

	for element in moreElements:

		if len(course_string_input) == len(element.text):
			element.click()
			break

	time.sleep(2)

	searchBox.send_keys(u"\ue007")

def get_links_from_course(driver):

	elements = driver.find_elements_by_tag_name("a")

	links = []

	for element in elements:
		if element.get_attribute("class") == "poll record":
			links.append(element.get_attribute("href"))

	links = list(dict.fromkeys(links))

	return links

def get_course_students_oed(driver, observer_link, course_name, year, term_name):

	click_course(driver, observer_link, course_name)
	time.sleep(2)

	links = get_links_from_course(driver)

	students = []

	for link in links:
		current_student = Student()
		current_student.import_oed_info(driver, link, year, term_name)
		students.append(current_student)
	
	return students

def get_async_course_students_oed(driver, observer_link, course_name, year, term_name):

	click_course(driver, observer_link, course_name)
	time.sleep(2)

	links = get_links_from_course(driver)

	tasks = []
	loop = asyncio.get_event_loop()

	for link in links:

		group_creator = loop.run_in_executor(None, create_student, driver, link, year, term_name)
		tasks.append(group_creator)
	
	students = loop.run_until_complete(asyncio.gather(*tasks))
	
	return students


def create_student(driver, link, year, term):

	current_student = Student()
	driver_copy = copy.copy(driver)
	current_student.import_oed_info(driver_copy, link, year, term)
	
	return current_student
	

def update_students_ed_info(driver, observer_link, course_name, student_list, year, term_number):

	click_course(driver, observer_link, course_name)
	time.sleep(2)

	links = get_links_from_course(driver)

	for link in links:

		driver.get(link)
		student_info = driver.find_elements_by_css_selector("span[person_role=\"estudiante\"]")[0]
		student_id = int(student_info.get_attribute("id").split('_')[2])

		for student in student_list:
			if student.id == student_id:
				student.import_ed_info(driver, link, year, term_number)

def write_to_excel(driver, course_link, course_name, student_list, term):
	
	driver.get(course_link)
	course_words = course_name.split(' ')
	course_base_name = unidecode.unidecode(course_words[0].strip())
	course_number = course_words[-1]

	course_box_selector = f"select[id=\"sections_in_3\"]"
	course_box = driver.find_element_by_css_selector(course_box_selector)
	courses = course_box.find_elements_by_css_selector('optgroup')
	
	for element in courses:
		element_label = unidecode.unidecode(element.get_attribute('label')).upper().strip()

		if element_label == course_base_name:
			options = element.find_elements_by_css_selector('*')
			
			for option in options:

				if option.text  == course_number:
					option.click()
	
	submit_button_selector = "input[name=\"submit\"]"
	submit_button = driver.find_element_by_css_selector(submit_button_selector)

	submit_button.click()
	
	course_info_elements = driver.find_elements_by_css_selector("span[class=\"value\"]")

	course_info = [element.text for element in course_info_elements]

	course_year = course_info[0]
	course_front_name = course_info[1]
	course_leader = course_info[2].upper()
	
	course_wb = openpyxl.load_workbook('Preset.xlsx')
	ws = course_wb.active

	ws.cell(2, 3).value = course_year
	ws.cell(3, 3).value = course_front_name
	ws.cell(4, 3).value = course_leader

	base_term_column = 4
	term_column = base_term_column + term - 1

	base_student_row = 7
	base_student_column = 2

	for i, student in enumerate(student_list):

		student_row = base_student_row + i
		number_column = base_student_column
		name_column = base_student_column + 1
		observation_column = 9

		ws.cell(student_row, number_column).value = i + 1
		ws.cell(student_row, name_column).value = str(student)
		ws.cell(student_row, term_column).value = student.grade

		oed_observations = [f'OED {observation}' for observation in student.oed_observations]
		ed_observations = [f'ED {observation}' for observation in student.ed_observations]
		observations = ' '.join(oed_observations + ed_observations)

		ws.cell(student_row, observation_column).value = observations

	course_wb.save(f'{course_front_name}.xlsx')

def main(headless = False):
	
	username = "william.parada81"
	password = "Caballeria1972" 

	if headless:

		options = Options()
		options.headless = True
		driver = webdriver.Chrome(WD_PATH, options=options)
	
	else:
		driver = webdriver.Chrome(WD_PATH)
		print("Usando Chrome con cabeza")

	driver.get("https://colmilgeneralsantander.phidias.co/")

	userBox = driver.find_element_by_id("autofocus")
	userBox.send_keys(username)

	passwordBox = driver.find_elements_by_xpath("//*[@id=\"login_form\"]/fieldset/div[2]/input")[0]
	passwordBox.send_keys(password)

	passwordBox.send_keys(u"\ue007")

	courses = {}

	courses["PREJARDIN"] = ["PREJARDIN 1"]
	courses["JARDIN"] = ["JARDIN 1"]
	courses["TRANSICION"] = ["TRANSICION 1"]
	courses["PRIMERO"] = ["PRIMERO 1"]
	courses["SEGUNDO"] = ["SEGUNDO 1"]
	courses["TERCERO"] = ["TERCERO 1"]
	# courses["CUARTO"] = ["CUARTO 1"]
	# courses["QUINTO"] = ["QUINTO 1"]
	# courses["SEXTO"] = ["SEXTO 1", "SEXTO 2"]
	# courses["SEPTIMO"] = ["SEPTIMO 1", "SEPTIMO 2"]
	# courses["OCTAVO"] = ["OCTAVO 1", "OCTAVO 2", "OCTAVO 3"]
	# courses["NOVENO"] = ["NOVENO 1", "NOVENO 2"]
	# courses["DECIMO"] = ["DECIMO 1", "DECIMO 2"]
	# courses["UNDECIMO"] = ["UNDECIMO 1"]

	for courseGrade in courses:

		for course in courses[courseGrade]:

			courseName = F"{courseGrade} {course}"

			students = get_course_students_oed(driver, OED_PAGE, courseName, YEAR, TERM_NAME)
			update_students_ed_info(driver, ED_PAGE, courseName, students, YEAR, TERM_NUMBER)

			write_to_excel(driver, 'https://colmilgeneralsantander.phidias.co/person/printables/checklist', courseName, students, TERM_NUMBER)
			

main()
