import os
import openpyxl
from Persona import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors

EXCEL_PAGE = "input.xlsx"

def main():
    print_title()

    wb = openpyxl.load_workbook(EXCEL_PAGE)
    ws = wb.active

    people = []
    documents = []

    h = 0
    n = 0

    for row in ws.iter_rows(2, ws.max_row, 1, 15):
        doc = row[0].value

        if doc is None:
            n += 1
            continue

        if doc not in documents:
            people.append(Persona(doc, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[8].value))
            documents.append(doc)
        
        person_index = search_person(people, doc)
        people[person_index].add_event(row[7].value, row[9].value, row[10].value, row[11].value, row[12].value, row[13].value, row[14].value)
        h += 1

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    for row in ws.iter_rows(1,1,1,15):
        for cell in row:
            output_ws.cell(cell.row, cell.column).value = cell.value
            output_ws.cell(cell.row, cell.column).font = Font(name='Consolas', bold=True)
            output_ws.cell(cell.row, cell.column).alignment = Alignment(horizontal='center')

    row_number = 2

    for person in people:
        fill = False
        
        if not person.get_events_validity():
            fill = True
        
        for event in person.events:
            output_ws.cell(row_number, 1).value = person.document
            output_ws.cell(row_number, 2).value = person.sur1
            output_ws.cell(row_number, 3).value = person.sur2
            output_ws.cell(row_number, 4).value = person.name1
            output_ws.cell(row_number, 5).value = person.name2
            output_ws.cell(row_number, 6).value = person.genre
            output_ws.cell(row_number, 7).value = person.bday
            output_ws.cell(row_number, 8).value = event[0]
            output_ws.cell(row_number, 9).value = person.code
            output_ws.cell(row_number, 10).value = event[1]
            output_ws.cell(row_number, 11).value = event[2]
            output_ws.cell(row_number, 12).value = event[3]
            output_ws.cell(row_number, 13).value = event[4]
            output_ws.cell(row_number, 14).value = event[5]
            output_ws.cell(row_number, 15).value = event[6]

            for row in output_ws.iter_rows(row_number, row_number, 1, 15):
                    for cell in row:
                        cell.font = Font(name='Consolas')
                        cell.alignment = Alignment(horizontal='center')

            if fill:
                redFill = PatternFill(start_color='FFFF0000',
                end_color='FFFF0000',
                fill_type='solid')
                for row in output_ws.iter_rows(row_number, row_number, 1, 15):
                    for cell in row:
                        cell.fill = redFill

            row_number += 1

    output_wb.save("Output.xlsx")



        
def print_title():
    print(r"""  ____     U  ___ u   ____      _____  U _____ u   ____     
 / __"| u   \/"_ \/U |  _"\ u  |_ " _| \| ___"|/U |  _"\ u  
<\___ \/    | | | | \| |_) |/    | |    |  _|"   \| |_) |/  
 u___) |.-,_| |_| |  |  _ <     /| |\   | |___    |  _ <    
 |____/>>\_)-\___/   |_| \_\   u |_|U   |_____|   |_| \_\   
  )(  (__)    \\     //   \\_  _// \\_  <<   >>   //   \\_  
 (__)        (__)   (__)  (__)(__) (__)(__) (__) (__)  (__) """)

def search_person(people, document):

    i = 0
    for person in people:
        if person.document == document:
            return i
        
        i += 1

if __name__ == "__main__":
    main()